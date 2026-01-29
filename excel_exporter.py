
"""
Excel export functionality
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

def export_to_excel(df, budget_data, risk_data, optimizations, filepath):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Budget Data"
        
        ws['A1'] = f"Budget Analysis - {budget_data['filename']}"
        ws['A1'].font = Font(size=16, bold=True)
        
        ws['A3'] = f"Total Budget: ${budget_data['total_budget']:,.2f}"
        ws['A4'] = f"Line Items: {budget_data['line_items']}"
        
        if not df.empty:
            start_row = 6
            headers = list(df.columns)
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=start_row, column=col_idx, value=header).font = Font(bold=True)
            
            for row_idx, (_, row) in enumerate(df.iterrows(), start_row + 1):
                for col_idx, col_name in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row[col_name])
        
        wb.save(filepath)
        print(f"✅ Excel saved: {filepath}")
        return True
    except Exception as e:
        print(f"❌ Excel error: {e}")
        return False

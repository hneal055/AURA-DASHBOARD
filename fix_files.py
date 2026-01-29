import os

print("🛠️ FIXING ALL MISSING AND INVALID FILES...")
print("=" * 50)

# Content for risk_manager.py
risk_manager_content = '''
"""
Risk Management Module for Budget Analysis
AI-powered risk detection and assessment
"""

import pandas as pd
from datetime import datetime

class RiskManager:
    def __init__(self):
        self.risk_categories = {
            'high_cost_items': {
                'threshold': 0.1,
                'severity': 'HIGH',
                'description': 'Items costing more than 10% of total budget'
            }
        }
    
    def analyze_risks(self, df):
        total_budget = df['Amount'].sum()
        risk_items = {
            'high_cost_items': self._detect_high_cost_items(df, total_budget)
        }
        
        overall_score = self._calculate_overall_risk_score(risk_items)
        risk_level = self._get_risk_level(overall_score)
        
        return {
            'summary': {
                'overall_risk_score': overall_score,
                'risk_level': risk_level,
                'total_risks_found': sum(len(items) for items in risk_items.values())
            },
            'items_by_category': risk_items,
            'recommendations': []
        }
    
    def _detect_high_cost_items(self, df, total_budget):
        threshold = self.risk_categories['high_cost_items']['threshold']
        high_cost_threshold = total_budget * threshold
        high_cost_items = df[df['Amount'] > high_cost_threshold].copy()
        
        risks = []
        for _, item in high_cost_items.iterrows():
            percentage = (item['Amount'] / total_budget * 100) if total_budget > 0 else 0
            risks.append({
                'description': item.get('Description', 'Unknown'),
                'amount': item['Amount'],
                'percentage_of_total': percentage,
                'department': item.get('Department', 'Unknown')
            })
        return risks
    
    def _calculate_overall_risk_score(self, risk_items):
        total_risks = sum(len(items) for items in risk_items.values())
        return min(100, total_risks * 10)
    
    def _get_risk_level(self, score):
        if score >= 70: return 'HIGH'
        elif score >= 40: return 'MEDIUM'
        elif score >= 20: return 'LOW'
        else: return 'VERY_LOW'
'''

# Content for charts_data.py
charts_data_content = '''
"""
Chart data preparation and HTML generation
"""

import json

def prepare_chart_data(df):
    chart_data = {}
    try:
        if 'Department' in df.columns:
            dept_data = df.groupby('Department')['Amount'].sum().sort_values(ascending=False)
            chart_data['department_breakdown'] = {
                'labels': dept_data.index.tolist(),
                'data': dept_data.values.tolist(),
                'colors': ['#3498db', '#2ecc71', '#e74c3c', '#f39c12']
            }
    except Exception as e:
        print(f"Chart data error: {e}")
    return chart_data

def generate_chart_html(chart_data):
    if not chart_data:
        return "<div>No chart data available</div>"
    
    html = """
    <div class="card fade-in">
        <h2 class="card-title">📊 Budget Visualizations</h2>
        <div class="charts-grid">
    """
    
    if 'department_breakdown' in chart_data:
        dept_data = chart_data['department_breakdown']
        html += f"""
            <div class="chart-container">
                <h3>🏢 Department Allocation</h3>
                <canvas id="deptChart" width="400" height="300"></canvas>
                <script>
                    setTimeout(() => {{
                        new Chart(document.getElementById('deptChart'), {{
                            type: 'doughnut',
                            data: {{
                                labels: {json.dumps(dept_data['labels'])},
                                datasets: [{{
                                    data: {json.dumps(dept_data['data'])},
                                    backgroundColor: {json.dumps(dept_data['colors'])}
                                }}]
                            }},
                            options: {{ responsive: true }}
                        }});
                    }}, 100);
                </script>
            </div>
        """
    
    html += """
        </div>
    </div>
    """
    return html
'''

# Content for excel_exporter.py
excel_exporter_content = '''
"""
Excel export functionality
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font

def export_to_excel(df, budget_data, risk_data, optimizations, filepath):
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Budget Data"
        
        ws['A1'] = f"Budget Analysis - {budget_data['filename']}"
        ws['A1'].font = Font(size=16, bold=True)
        
        ws['A3'] = f"Total Budget: ${budget_data['total_budget']:,.2f}"
        ws['A4'] = f"Line Items: {budget_data['line_items']}"
        
        if not df.empty:
            start_row = 6
            headers = list(df.columns)
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=start_row, column=col_idx, value=header).font = Font(bold=True)
            
            for row_idx, (_, row) in enumerate(df.iterrows(), start_row + 1):
                for col_idx, col_name in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row[col_name])
        
        wb.save(filepath)
        print(f"✅ Excel saved: {filepath}")
        return True
    except Exception as e:
        print(f"❌ Excel error: {e}")
        return False
'''

# Content for budget_comparison.py
budget_comparison_content = '''
"""
Budget comparison functionality
"""

import pandas as pd

def compare_budgets(df1, df2, name1="Budget 1", name2="Budget 2"):
    total1 = df1['Amount'].sum()
    total2 = df2['Amount'].sum()
    total_change = total2 - total1
    percent_change = (total_change / total1 * 100) if total1 != 0 else 0
    
    dept_changes = {}
    if 'Department' in df1.columns and 'Department' in df2.columns:
        all_depts = set(df1['Department'].unique()) | set(df2['Department'].unique())
        for dept in all_depts:
            amount1 = df1[df1['Department'] == dept]['Amount'].sum()
            amount2 = df2[df2['Department'] == dept]['Amount'].sum()
            difference = amount2 - amount1
            dept_changes[dept] = {
                'budget1_amount': amount1,
                'budget2_amount': amount2,
                'difference': difference,
                'percent_change': (difference / amount1 * 100) if amount1 != 0 else 0,
                'status': 'increased' if difference > 0 else 'decreased' if difference < 0 else 'unchanged'
            }
    
    insights = []
    if total_change > 0:
        insights.append(f"Budget increased by ${abs(total_change):,.0f}")
    elif total_change < 0:
        insights.append(f"Budget decreased by ${abs(total_change):,.0f}")
    
    return {
        'budget1_name': name1,
        'budget2_name': name2,
        'budget1_total': total1,
        'budget2_total': total2,
        'total_change': total_change,
        'percent_change': percent_change,
        'department_changes': dept_changes,
        'insights': insights
    }
'''

# Content for comparison_charts.py
comparison_charts_content = '''
"""
Comparison chart generation
"""

import json

def generate_comparison_chart_html(comparison_data):
    if not comparison_data:
        return ""
    
    return f"""
    <div class="card fade-in">
        <h2 class="card-title">📊 Comparison Visualizations</h2>
        <div class="chart-container">
            <h3>💰 Overall Budget Change</h3>
            <canvas id="overallComparisonChart" width="400" height="400"></canvas>
            <script>
                new Chart(document.getElementById('overallComparisonChart'), {{
                    type: 'bar',
                    data: {{
                        labels: ['{comparison_data['budget1_name']}', '{comparison_data['budget2_name']}'],
                        datasets: [{{
                            label: 'Total Budget',
                            data: [{comparison_data['budget1_total']}, {comparison_data['budget2_total']}],
                            backgroundColor: ['rgba(52, 152, 219, 0.7)', 'rgba(46, 204, 113, 0.7)']
                        }}]
                    }},
                    options: {{ responsive: true }}
                }});
            </script>
        </div>
    </div>
    """
'''

# Content for budget_templates.py
budget_templates_content = '''
"""
Budget Templates Database
"""

BUDGET_TEMPLATES = {
    "film_production": {
        "id": "film_production",
        "name": "🎬 Film Production Budget",
        "category": "Creative & Media",
        "description": "Complete budget for film/video production",
        "icon": "🎬",
        "popularity": 95,
        "line_items": [
            {"Description": "Script Development", "Department": "Pre-Production", "Category": "Creative", "Amount": 5000},
            {"Description": "Director Fee", "Department": "Production", "Category": "Talent", "Amount": 15000},
            {"Description": "Camera Equipment", "Department": "Production", "Category": "Equipment", "Amount": 8000}
        ],
        "total_amount": 28000,
        "item_count": 3,
        "departments": ["Pre-Production", "Production"]
    },
    "marketing_campaign": {
        "id": "marketing_campaign", 
        "name": "📢 Marketing Campaign Budget",
        "category": "Business & Marketing",
        "description": "Budget for digital marketing",
        "icon": "📢",
        "popularity": 88,
        "line_items": [
            {"Description": "Social Media Ads", "Department": "Digital", "Category": "Advertising", "Amount": 10000},
            {"Description": "Content Creation", "Department": "Content", "Category": "Creative", "Amount": 8000}
        ],
        "total_amount": 18000,
        "item_count": 2,
        "departments": ["Digital", "Content"]
    }
}

def get_template_categories():
    categories = {}
    for template in BUDGET_TEMPLATES.values():
        category = template['category']
        if category not in categories:
            categories[category] = []
        categories[category].append(template)
    return categories

def get_popular_templates(limit=6):
    return list(BUDGET_TEMPLATES.values())[:limit]

def get_template_by_id(template_id):
    return BUDGET_TEMPLATES.get(template_id)

def search_templates(query):
    query = query.lower()
    results = []
    for template in BUDGET_TEMPLATES.values():
        if (query in template['name'].lower() or 
            query in template['description'].lower()):
            results.append(template)
    return results
'''

# Create all files
files_to_create = {
    'risk_manager.py': risk_manager_content,
    'charts_data.py': charts_data_content,
    'excel_exporter.py': excel_exporter_content,
    'budget_comparison.py': budget_comparison_content,
    'comparison_charts.py': comparison_charts_content,
    'budget_templates.py': budget_templates_content
}

for filename, content in files_to_create.items():
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(content)
        print(f"✅ CREATED: {filename}")
    except Exception as e:
        print(f"❌ FAILED to create {filename}: {e}")

print("\\n" + "=" * 50)
print("🎉 ALL FILES HAVE BEEN CREATED/FIXED!")
print("Now run: python check_file_contents.py")
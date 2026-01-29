import os

print("🛠️ FIXING ALL MISSING AND INVALID FILES...")
print("=" * 50)

# ============================================================================
# 1. FIX risk_manager.py (replace invalid content)
# ============================================================================
risk_manager_content = '''"""
Risk Management Module for Budget Analysis
AI-powered risk detection and assessment
"""

import pandas as pd
import numpy as np
from datetime import datetime

class RiskManager:
    def __init__(self):
        self.risk_categories = {
            'high_cost_items': {
                'threshold': 0.1,
                'severity': 'HIGH',
                'description': 'Items costing more than 10% of total budget'
            },
            'vendor_concentration': {
                'threshold': 0.3,
                'severity': 'MEDIUM', 
                'description': 'High spending concentration with single vendor'
            }
        }
    
    def analyze_risks(self, df):
        """
        Comprehensive risk analysis of budget data
        """
        total_budget = df['Amount'].sum()
        
        # Simple risk analysis
        risk_items = {
            'high_cost_items': self._detect_high_cost_items(df, total_budget),
            'vendor_concentration': self._detect_vendor_concentration(df, total_budget)
        }
        
        # Calculate overall risk score
        overall_score = self._calculate_overall_risk_score(risk_items)
        risk_level = self._get_risk_level(overall_score)
        
        return {
            'summary': {
                'overall_risk_score': overall_score,
                'risk_level': risk_level,
                'total_risks_found': sum(len(items) for items in risk_items.values()),
                'analysis_timestamp': datetime.now().isoformat()
            },
            'items_by_category': risk_items,
            'recommendations': []
        }
    
    def _detect_high_cost_items(self, df, total_budget):
        """Detect items costing more than threshold percentage of total budget"""
        threshold = self.risk_categories['high_cost_items']['threshold']
        high_cost_threshold = total_budget * threshold
        
        high_cost_items = df[df['Amount'] > high_cost_threshold].copy()
        
        risks = []
        for _, item in high_cost_items.iterrows():
            percentage = (item['Amount'] / total_budget * 100) if total_budget > 0 else 0
            risks.append({
                'description': item.get('Description', 'Unknown'),
                'amount': item['Amount'],
                'percentage_of_total': percentage,
                'department': item.get('Department', 'Unknown'),
                'risk_reason': f"Costs {percentage:.1f}% of total budget"
            })
        
        return risks
    
    def _detect_vendor_concentration(self, df, total_budget):
        """Detect high spending concentration with single vendors"""
        if 'Vendor' not in df.columns:
            return []
            
        vendor_totals = df.groupby('Vendor')['Amount'].sum()
        threshold = self.risk_categories['vendor_concentration']['threshold']
        
        risks = []
        for vendor, amount in vendor_totals.items():
            percentage = (amount / total_budget * 100) if total_budget > 0 else 0
            if percentage > threshold * 100:
                risks.append({
                    'vendor': vendor,
                    'amount': amount,
                    'percentage_of_total': percentage,
                    'risk_reason': f"Concentrates {percentage:.1f}% of budget with one vendor"
                })
        
        return risks
    
    def _calculate_overall_risk_score(self, risk_items):
        """Calculate overall risk score from 0-100"""
        total_risks = sum(len(items) for items in risk_items.values())
        return min(100, total_risks * 10)  # Simple scoring
    
    def _get_risk_level(self, score):
        """Convert numerical score to risk level"""
        if score >= 70:
            return 'HIGH'
        elif score >= 40:
            return 'MEDIUM'
        elif score >= 20:
            return 'LOW'
        else:
            return 'VERY_LOW'
'''

# ============================================================================
# 2. CREATE charts_data.py (missing file)
# ============================================================================
charts_data_content = '''"""
Chart data preparation and HTML generation
"""

import json

def prepare_chart_data(df):
    """Prepare data for chart visualizations"""
    chart_data = {}
    
    try:
        # Department breakdown
        if 'Department' in df.columns:
            dept_data = df.groupby('Department')['Amount'].sum().sort_values(ascending=False)
            chart_data['department_breakdown'] = {
                'labels': dept_data.index.tolist(),
                'data': dept_data.values.tolist(),
                'colors': ['#3498db', '#2ecc71', '#e74c3c', '#f39c12', '#9b59b6']
            }
        
        # Category breakdown  
        if 'Category' in df.columns:
            cat_data = df.groupby('Category')['Amount'].sum().sort_values(ascending=False)
            chart_data['category_breakdown'] = {
                'labels': cat_data.index.tolist(),
                'data': cat_data.values.tolist()
            }
            
    except Exception as e:
        print(f"Chart data preparation error: {e}")
    
    return chart_data

def generate_chart_html(chart_data):
    """Generate HTML for charts"""
    if not chart_data:
        return """
        <div class="card fade-in">
            <h2 class="card-title">📊 Budget Visualizations</h2>
            <p style="text-align: center; color: #7f8c8d; padding: 40px;">
                No chart data available for visualization
            </p>
        </div>
        """
    
    html = """
    <div class="card fade-in">
        <div class="card-header">
            <h2 class="card-title">📊 Budget Visualizations</h2>
            <p class="card-subtitle">Interactive charts for better budget insights</p>
        </div>
        <div class="charts-grid">
    """
    
    # Department Chart
    if 'department_breakdown' in chart_data:
        dept_data = chart_data['department_breakdown']
        html += f"""
            <div class="chart-container">
                <h3>🏢 Department Allocation</h3>
                <canvas id="deptChart" width="400" height="300"></canvas>
                <script>
                    setTimeout(() => {{
                        new Chart(document.getElementById('deptChart'), {{
                            type: 'doughnut',
                            data: {{
                                labels: {json.dumps(dept_data['labels'])},
                                datasets: [{{
                                    data: {json.dumps(dept_data['data'])},
                                    backgroundColor: {json.dumps(dept_data.get('colors', ['#3498db', '#2ecc71', '#e74c3c', '#f39c12', '#9b59b6']))},
                                    borderWidth: 2,
                                    borderColor: '#fff'
                                }}]
                            }},
                            options: {{
                                responsive: true,
                                maintainAspectRatio: false,
                                plugins: {{
                                    legend: {{ 
                                        position: 'right',
                                        labels: {{ usePointStyle: true, padding: 20 }}
                                    }},
                                    tooltip: {{
                                        callbacks: {{
                                            label: function(context) {{
                                                const label = context.label || '';
                                                const value = context.parsed;
                                                const total = context.dataset.data.reduce((a, b) => a + b, 0);
                                                const percentage = Math.round((value / total) * 100);
                                                return `${{label}}: ${{value.toLocaleString()}} (${{percentage}}%)`;
                                            }}
                                        }}
                                    }}
                                }}
                            }}
                        }});
                    }}, 100);
                </script>
            </div>
        """
    
    html += """
        </div>
    </div>
    
    <style>
    .charts-grid {
        display: grid;
        grid-template-columns: repeat(auto-fit, minmax(400px, 1fr));
        gap: 20px;
        margin-top: 20px;
    }
    .chart-container {
        background: white;
        padding: 20px;
        border-radius: 8px;
        box-shadow: 0 2px 4px rgba(0,0,0,0.1);
    }
    .chart-container h3 {
        margin: 0 0 15px 0;
        color: #2c3e50;
    }
    </style>
    """
    
    return html
'''

# ============================================================================
# 3. CREATE excel_exporter.py (missing file)
# ============================================================================
excel_exporter_content = '''"""
Excel export functionality for budget analysis
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def export_to_excel(df, budget_data, risk_data, optimizations, filepath):
    """Export budget analysis to formatted Excel file"""
    try:
        wb = Workbook()
        
        # Budget Data Sheet
        ws_budget = wb.active
        ws_budget.title = "Budget Data"
        
        # Header
        ws_budget['A1'] = f"Budget Analysis - {budget_data['filename']}"
        ws_budget['A1'].font = Font(size=16, bold=True)
        
        # Summary info
        ws_budget['A3'] = f"Total Budget: ${budget_data['total_budget']:,.2f}"
        ws_budget['A4'] = f"Line Items: {budget_data['line_items']}"
        ws_budget['A5'] = f"Risk Level: {risk_data.get('risk_level', 'N/A')}"
        
        # Data table
        if not df.empty:
            start_row = 7
            
            # Write column headers
            headers = list(df.columns)
            for col_idx, header in enumerate(headers, 1):
                cell = ws_budget.cell(row=start_row, column=col_idx, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
            
            # Write data rows
            for row_idx, (_, row) in enumerate(df.iterrows(), start_row + 1):
                for col_idx, col_name in enumerate(headers, 1):
                    value = row[col_name]
                    ws_budget.cell(row=row_idx, column=col_idx, value=value)
        
        # Save the file
        wb.save(filepath)
        print(f"✅ Excel file saved: {filepath}")
        return True
        
    except Exception as e:
        print(f"❌ Excel export error: {e}")
        return False
'''

# ============================================================================
# 4. CREATE budget_comparison.py (missing file)
# ============================================================================
budget_comparison_content = '''"""
Budget comparison functionality
"""

import pandas as pd

def compare_budgets(df1, df2, name1="Budget 1", name2="Budget 2"):
    """
    Compare two budgets and return detailed analysis
    """
    
    # Basic metrics
    total1 = df1['Amount'].sum()
    total2 = df2['Amount'].sum()
    items1 = len(df1)
    items2 = len(df2)
    
    # Calculate changes
    total_change = total2 - total1
    percent_change = (total_change / total1 * 100) if total1 != 0 else 0
    
    # Department analysis
    dept_changes = {}
    if 'Department' in df1.columns and 'Department' in df2.columns:
        all_depts = set(df1['Department'].unique()) | set(df2['Department'].unique())
        
        for dept in all_depts:
            dept1 = df1[df1['Department'] == dept]
            dept2 = df2[df2['Department'] == dept]
            
            amount1 = dept1['Amount'].sum() if not dept1.empty else 0
            amount2 = dept2['Amount'].sum() if not dept2.empty else 0
            
            difference = amount2 - amount1
            percent_change_dept = (difference / amount1 * 100) if amount1 != 0 else 0
            
            # Determine status
            if amount1 == 0 and amount2 > 0:
                status = "new"
            elif amount1 > 0 and amount2 == 0:
                status = "removed"
            elif difference > 0:
                status = "increased"
            elif difference < 0:
                status = "decreased"
            else:
                status = "unchanged"
            
            dept_changes[dept] = {
                'budget1_amount': amount1,
                'budget2_amount': amount2,
                'difference': difference,
                'percent_change': percent_change_dept,
                'status': status
            }
    
    # Generate insights
    insights = []
    if total_change > 0:
        insights.append(f"Overall budget increased by ${abs(total_change):,.0f} ({percent_change:.1f}%)")
    elif total_change < 0:
        insights.append(f"Overall budget decreased by ${abs(total_change):,.0f} ({abs(percent_change):.1f}%)")
    else:
        insights.append("Overall budget remained unchanged")
    
    return {
        'budget1_name': name1,
        'budget2_name': name2,
        'budget1_total': total1,
        'budget2_total': total2,
        'budget1_items': items1,
        'budget2_items': items2,
        'total_change': total_change,
        'percent_change': percent_change,
        'department_changes': dept_changes,
        'new_items': [],
        'removed_items': [],
        'total_new_items': 0,
        'total_removed_items': 0,
        'insights': insights
    }
'''

# ============================================================================
# 5. CREATE comparison_charts.py (missing file)
# ============================================================================
comparison_charts_content = '''"""
Comparison chart generation
"""

import json

def generate_comparison_chart_html(comparison_data):
    """Generate HTML for comparison charts"""
    if not comparison_data:
        return ""
    
    html = """
    <div class="card fade-in">
        <div class="card-header">
            <h2 class="card-title">📊 Comparison Visualizations</h2>
            <p class="card-subtitle">Side-by-side analysis of budget changes</p>
        </div>
        <div class="charts-grid">
    """
    
    # Overall budget comparison
    html += f"""
        <div class="chart-container">
            <h3>💰 Overall Budget Change</h3>
            <canvas id="overallComparisonChart" width="400" height="400"></canvas>
            <script>
                setTimeout(() => {{
                    new Chart(document.getElementById('overallComparisonChart'), {{
                        type: 'bar',
                        data: {{
                            labels: ['{comparison_data['budget1_name']}', '{comparison_data['budget2_name']}'],
                            datasets: [{{
                                label: 'Total Budget',
                                data: [{comparison_data['budget1_total']}, {comparison_data['budget2_total']}],
                                backgroundColor: [
                                    'rgba(52, 152, 219, 0.7)',
                                    'rgba(46, 204, 113, 0.7)'
                                ],
                                borderColor: [
                                    'rgba(52, 152, 219, 1)',
                                    'rgba(46, 204, 113, 1)'
                                ],
                                borderWidth: 2
                            }}]
                        }},
                        options: {{
                            responsive: true,
                            maintainAspectRatio: false,
                            scales: {{
                                y: {{
                                    beginAtZero: true,
                                    ticks: {{
                                        callback: function(value) {{
                                            return '$' + value.toLocaleString();
                                        }}
                                    }}
                                }}
                            }},
                            plugins: {{
                                legend: {{ display: false }},
                                tooltip: {{
                                    callbacks: {{
                                        label: function(context) {{
                                            return '$' + context.parsed.y.toLocaleString();
                                        }}
                                    }}
                                }}
                            }}
                        }}
                    }});
                }}, 200);
            </script>
        </div>
    """
    
    html += """
        </div>
    </div>
    """
    
    return html
'''

# ============================================================================
# 6. FIX budget_templates.py (replace invalid content)
# ============================================================================
budget_templates_content = '''"""
Budget Templates Database
Pre-built templates for common industries and use cases
"""

BUDGET_TEMPLATES = {
    "film_production": {
        "id": "film_production",
        "name": "🎬 Film Production Budget",
        "category": "Creative & Media",
        "description": "Complete budget for film/video production",
        "icon": "🎬",
        "popularity": 95,
        "line_items": [
            {"Description": "Script Development", "Department": "Pre-Production", "Category": "Creative", "Amount": 5000},
            {"Description": "Director Fee", "Department": "Production", "Category": "Talent", "Amount": 15000},
            {"Description": "Camera Equipment", "Department": "Production", "Category": "Equipment", "Amount": 8000},
            {"Description": "Video Editing", "Department": "Post-Production", "Category": "Creative", "Amount": 8000}
        ],
        "total_amount": 36000,
        "item_count": 4,
        "departments": ["Pre-Production", "Production", "Post-Production"]
    },
    "marketing_campaign": {
        "id": "marketing_campaign", 
        "name": "📢 Marketing Campaign Budget",
        "category": "Business & Marketing",
        "description": "Budget for digital and traditional marketing",
        "icon": "📢",
        "popularity": 88,
        "line_items": [
            {"Description": "Social Media Ads", "Department": "Digital", "Category": "Advertising", "Amount": 10000},
            {"Description": "Content Creation", "Department": "Content", "Category": "Creative", "Amount": 8000},
            {"Description": "Google Ads", "Department": "Digital", "Category": "Advertising", "Amount": 7000}
        ],
        "total_amount": 25000,
        "item_count": 3,
        "departments": ["Digital", "Content"]
    }
}

def get_template_categories():
    """Get all unique template categories"""
    categories = {}
    for template in BUDGET_TEMPLATES.values():
        category = template['category']
        if category not in categories:
            categories[category] = []
        categories[category].append(template)
    return categories

def get_popular_templates(limit=6):
    """Get most popular templates"""
    return list(BUDGET_TEMPLATES.values())[:limit]

def get_template_by_id(template_id):
    """Get a specific template by ID"""
    return BUDGET_TEMPLATES.get(template_id)

def search_templates(query):
    """Search templates by name, description, or category"""
    query = query.lower()
    results = []
    for template in BUDGET_TEMPLATES.values():
        if (query in template['name'].lower() or 
            query in template['description'].lower() or 
            query in template['category'].lower()):
            results.append(template)
    return results
'''

# ============================================================================
# CREATE ALL FILES
# ============================================================================
files_to_create = {
    'risk_manager.py': risk_manager_content,
    'charts_data.py': charts_data_content,
    'excel_exporter.py': excel_exporter_content,
    'budget_comparison.py': budget_comparison_content,
    'comparison_charts.py': comparison_charts_content,
    'budget_templates.py': budget_templates_content
}

for filename, content in files_to_create.items():
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(content)
        print(f"✅ CREATED: {filename}")
    except Exception as e:
        print(f"❌ FAILED to create {filename}: {e}")

print("\\n" + "=" * 50)
print("🎉 ALL FILES HAVE BEEN CREATED/FIXED!")
print("Now run: python check_file_contents.py")



